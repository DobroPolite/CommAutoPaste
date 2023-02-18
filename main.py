import time
from random import randrange, shuffle
import os

import openpyxl as ox
from selenium import webdriver
from selenium.webdriver.common.by import By


def Comments_loader(name_file) -> list[str]:
    current_dir = os.getcwd()
    path = f"{current_dir}\{name_file}.xlsx"
    wb_obj = ox.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=3)
        comment_text = cell_obj.value
        comments.append(comment_text)
    return comments


def Comment_setter(url: str, countof: int):
    shuffle(comments)
    login = ''
    password = ''
    with open('accounts.txt', 'r') as f:
        for x in range(0, countof):
            login = f.readline().strip()
            password = f.readline().strip()
            options = webdriver.ChromeOptions()
            options.add_argument('headless')
            driver = webdriver.Chrome(options=options)
            driver.maximize_window()
            try:
                driver.get(url=url)
                driver.find_element(By.XPATH, '/html/body/div[1]/header/div[1]/div/div[3]/ul/li[2]/a').click()
                driver.find_element(By.NAME, 'username').send_keys(login)
                driver.find_element(By.ID, 'pass').send_keys(password)
                driver.find_element(By.XPATH, '//*[@id="login-form"]/div[3]/button').click()
                driver.find_element(By.XPATH, '//*[@id="c_comment_txt"]').send_keys(
                    comments[randrange(2, len(comments))])
                driver.find_element(By.XPATH, '//*[@id="c_submit"]').click()
                driver.close()
                driver.quit()
                print(f'Коммент готов! Осталось {countof-(x+1)} комментариев.')
            except Exception as ex:
                print(ex)


def main(name_file, url, countof):
    Comments_loader(name_file)
    Comment_setter(url, countof)


if __name__ == '__main__':
    a = 1
    while a != 0:
        try:
            comments = []
            name_file = str(input('Укажите только имя файла (файл должен быть в формате .xlsx): '))
            url = str(input('Введите ссылку на видео: '))
            countof = int(input('Введите количество комментариев: '))
            main(name_file, url, countof)
            a = a - 1
        except Exception as ex:
            print(ex)
            time.sleep(3)
            a = a - 1
else:
    print("Not __main file__")
